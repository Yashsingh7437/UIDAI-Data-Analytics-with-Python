import os
from dotenv import load_dotenv
import imaplib
import email
from email.header import decode_header
import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Color
from openpyxl.utils import get_column_letter

# Email credentials
load_dotenv()

username = os.getenv('EMAIL_USERNAME')
password = os.getenv('EMAIL_PASSWORD')
imap_server = os.getenv('IMAP_SERVER')  # e.g., 'imap.gmail.com' for Gmail

# Specific data to search in the Excel file
search_data = 'RO Lucknow'

# Connect to the server
mail = imaplib.IMAP4_SSL(imap_server)
mail.login(username, password)

# Select the mailbox you want to check
mail.select('inbox')

# Search for emails with the specific subject and sender
subject = "Here is the attachment"
sender_email = "devil10bro@gmail.com"
result, data = mail.search(None, f'(FROM "{sender_email}" SUBJECT "{subject}")')

# Fetch the email
email_ids = data[0].split()
if email_ids:
    for email_id in email_ids:
        result, message_data = mail.fetch(email_id, '(RFC822)')
        raw_email = message_data[0][1]
        msg = email.message_from_bytes(raw_email)

        # Check if the email has an attachment
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue

                # Check for the attachment and its name
                filename = part.get_filename()
                if filename:
                    # Decode the filename if necessary
                    filename = decode_header(filename)[0][0]
                    if isinstance(filename, bytes):
                        filename = filename.decode()

                    # Check if the file is an Excel file
                    if filename.endswith(('.xls', '.xlsx')):
                        # Load the Excel file into memory
                        attachment_data = part.get_payload(decode=True)
                        file_stream = BytesIO(attachment_data)
                        workbook = openpyxl.load_workbook(file_stream)
                        sheet = workbook.active

                        # Create a new workbook for the filtered data
                        new_workbook = openpyxl.Workbook()
                        new_sheet = new_workbook.active

                        # Function to apply styles to the first 4 rows
                        def style_first_rows(sheet, start_row, end_row):
                            for row_idx in range(start_row, end_row + 1):
                                for cell in sheet[row_idx]:
                                    cell.font = Font(bold=True)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Auto-adjust the column widths
                            for col_idx in range(1, sheet.max_column + 1):
                                col_letter = get_column_letter(col_idx)
                                max_length = 0
                                for row in range(1, end_row + 1):
                                    cell_value = new_sheet[f'{col_letter}{row}'].value
                                    if cell_value:
                                        max_length = max(max_length, len(str(cell_value)))
                                adjusted_width = max_length + 2
                                new_sheet.column_dimensions[col_letter].width = adjusted_width

                        # Copy the first 4 rows as they are
                        for row_idx in range(1, 5):
                            for col_idx, cell in enumerate(sheet[row_idx], 1):
                                new_cell = new_sheet.cell(row=row_idx, column=col_idx)
                                new_cell.value = cell.value

                        # Apply styles to the first 4 rows
                        style_first_rows(new_sheet, 1, 4)

                        # Copy merged cells in the first 4 rows
                        for merged_cell in sheet.merged_cells.ranges:
                            if merged_cell.min_row <= 4:
                                new_sheet.merge_cells(str(merged_cell))

                        # Search for the specific data in rows and copy the row if found
                        found = False
                        for row in sheet.iter_rows(2, sheet.max_row):
                            for cell in row:
                                if cell.value == search_data:
                                    new_row_idx = new_sheet.max_row + 1
                                    for col_idx, src_cell in enumerate(row, 1):
                                        dest_cell = new_sheet.cell(row=new_row_idx, column=col_idx)
                                        dest_cell.value = src_cell.value
        
                                    found = True
                                    break
                            if found:
                                break

                        # Save the new workbook
                        new_filename = 'filtered_data.xlsx'
                        new_workbook.save(new_filename)
                        print(f'Saved filtered data to {new_filename}')
                    else:
                        print("Attachment is there but it's not an Excel file")
else:
    print("No emails found with the specified subject and sender.")

# Logout and close the connection
mail.logout()
