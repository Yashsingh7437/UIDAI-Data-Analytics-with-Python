import os
from dotenv import load_dotenv
import imaplib
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime
import openpyxl
from io import BytesIO
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

# Load environment variables
load_dotenv()

username = os.getenv('EMAIL_USERNAME')
password = os.getenv('EMAIL_PASSWORD')
imap_server = os.getenv('IMAP_SERVER') 
sender_email = os.getenv('SENDER_EMAIL') 

# Specific data to search in the Excel file
search_data = 'RO Lucknow'

# Connect to the server
mail = imaplib.IMAP4_SSL(imap_server)
mail.login(username, password)

# Select the mailbox you want to check
mail.select('inbox')

# Search for emails with the specific subject and sender
subject = "Here is the attachment"
result, data = mail.search(None, f'(FROM "{sender_email}" SUBJECT "{subject}")')

# Fetch the email
email_ids = data[0].split()
if email_ids:
    for email_id in email_ids:
        result, message_data = mail.fetch(email_id, '(RFC822)')
        raw_email = message_data[0][1]
        msg = email.message_from_bytes(raw_email)

        # Extract the email date
        email_date = msg["Date"]
        email_date = parsedate_to_datetime(email_date).date()

        # Check if the email has an attachment
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue

                # Check for the attachment and its name
                filename = part.get_filename()
                if filename:
                    # Decode the filename if necessary
                    filename = decode_header(filename)[0][0]
                    if isinstance(filename, bytes):
                        filename = filename.decode()

                    # Check if the file is an Excel file
                    if filename.endswith(('.xls', '.xlsx')):
                        # Load the Excel file into memory
                        attachment_data = part.get_payload(decode=True)
                        file_stream = BytesIO(attachment_data)
                        workbook = openpyxl.load_workbook(file_stream)
                        sheet = workbook.active

                        # Load the existing Excel file to append data
                        existing_filename = 'existing_data.xlsx'
                        try:
                            existing_workbook = openpyxl.load_workbook(existing_filename)
                            existing_sheet = existing_workbook.active
                        except FileNotFoundError:
                            # If the file does not exist, create a new one
                            existing_workbook = openpyxl.Workbook()
                            existing_sheet = existing_workbook.active

                        # Function to check if two rows are identical
                        def rows_are_identical(row1, row2):
                            for cell1, cell2 in zip(row1, row2):
                                if cell1.value != cell2.value:
                                    return False
                            return True

                        # Function to delete identical rows
                        def delete_identical_rows(sheet):
                            rows_to_delete = []
                            for row_idx in range(2, sheet.max_row + 1):
                                row = sheet[row_idx]
                                for comp_row_idx in range(row_idx + 1, sheet.max_row + 1):
                                    comp_row = sheet[comp_row_idx]
                                    if rows_are_identical(row, comp_row):
                                        rows_to_delete.append(comp_row_idx)
                            for row_idx in sorted(rows_to_delete, reverse=True):
                                sheet.delete_rows(row_idx)

                        # Determine the starting row to append the new data
                        start_row = 5

                        # Shift existing data down if necessary
                        if existing_sheet.max_row >= start_row:
                            existing_sheet.insert_rows(start_row, amount=1)

                        # Search for the specific data in rows and copy the row if found
                        for row in sheet.iter_rows(2, sheet.max_row):
                            for cell in row:
                                if cell.value == search_data:
                                    # Shift the cells to the right to make space for the date
                                    for col_idx in range(sheet.max_column, 0, -1):
                                        existing_sheet.cell(row=start_row, column=col_idx+1).value = existing_sheet.cell(row=start_row, column=col_idx).value

                                    # Insert the email received date in the second column
                                    existing_sheet.cell(row=start_row, column=2).value = email_date.strftime('%Y-%m-%d')

                                    # Copy the rest of the data
                                    for col_idx, col in enumerate(row, 1):
                                        if col_idx >= 2:  # Skip the second column (date column)
                                            existing_sheet.cell(row=start_row, column=col_idx+1).value = col.value
                                    break

                        # Delete identical rows
                        delete_identical_rows(existing_sheet)

                        # Save the updated workbook
                        existing_workbook.save(existing_filename)
                        break 
                    else:
                        print("Attachment is there but it's not an Excel file")
                break 
    print(f'Saved updated data to {existing_filename}')
else:
    print("No emails found with the specified subject and sender.")

# Logout and close the connection
mail.logout()

# Load the updated Excel file for heatmap creation
workbook = openpyxl.load_workbook(existing_filename)
sheet = workbook.active

# Extract data starting from row 5
data = []
dates = []
for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=True):
    date_cell = row[1]  
    if date_cell:
        dates.append(date_cell)
    
    # Extract values from columns E and F, then calculate sum for column F
    val_e = row[4] if isinstance(row[4], (int, float)) else 0
    val_f = row[5] if isinstance(row[5], (int, float)) else 0
    val_g = row[6] if isinstance(row[6], (int, float)) else 0
    val_h = row[7] if isinstance(row[7], (int, float)) else 0
    
    # Update column F to be the sum of columns E and F
    data.append([val_e + val_f, val_g, val_h]) 

# Create a DataFrame
df = pd.DataFrame(data, columns=['F', 'G', 'H'])
df['Date'] = dates

# Set 'Date' as the index
df.set_index('Date', inplace=True)

# Create a copy of the DataFrame with updated column names for the heatmap
df_heatmap = df.rename(columns={'F': 'Total Cases', 'G': 'Closed Cases', 'H': 'Pending Cases'})

# Plot the heatmap
plt.figure(figsize=(8, 6))
sns.heatmap(df_heatmap.T, cmap='YlGnBu', annot=True, fmt=".1f")
plt.title('Heatmap of Total Cases, Closed Cases, Pending Cases')
plt.xlabel('Date')
plt.ylabel('Cases')
plt.xticks(rotation=45)
plt.show()
